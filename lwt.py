from openpyxl import Workbook, load_workbook
import re, datetime, sys, os, glob, platform, subprocess
from datetime import timedelta

"""
TASK: Search for unused variables
TASK: Put numbers after filenames for file selection
"""

class LWT:

    def __init__(self, filename=None, goals_filename=None):
        self.filename = filename
        self.workminutes = 'run'
        self.breakminutes = 'run'
        self.entry_date = ''
        self.entry_work_type = ''
        self.entry_work_notes = ''
        self.entry_work_duration = 0
        self.work_duration_position = ''
        self.work_notes_list = []
        self.project_goals_int_list = []
        #for max row in a sheet
        self.sheet_length = 0
        #turning minutes into hrs mins
        self.duration_hrs_mins = '0mins'
        self.work_alarm_file_path = ''
        self.break_alarm_file_path = ''

    """
    This method runs select() and alarm() to get the required files, then
    runs work(), because these are always the first thing I run
    """

    def start(self):

        self.select()
        self.alarm()
        self.work()

    """
    This creates a list of row numbers
    (self.project_goals_int_list) for goals that are
    listed in any of this project's entries.
    """
    def select(self):
        file_list = []
        project_list = []
        for file in glob.iglob('**/*.xlsx', recursive=True):
            file_list.append(file)
        for x in file_list:
            if re.search('tracking', x, re.IGNORECASE):
                project_list.append(x)
        for key, value in enumerate(project_list):
            print(str(key) + '. ' + value)
        print('\nenter the number for the tracking file\n')
        user_input = input()
        project_index = int(user_input)
        self.filename = project_list[project_index]

    """
    This presents all mp3 files and stores the selection in self.alarm_file_path
    """

    def alarm(self):
        file_list = []
        for file in glob.iglob('**/*.mp3', recursive=True):
            file_list.append(file)
        for key, value in enumerate(file_list):
            print(str(key) + '. ' + value)
        print('\nselect the work-timer alarm file by entering a number\n')
        user_input = input()
        file_index = int(user_input)
        self.work_alarm_file_path = file_list[file_index]
        print('\nselect break-timer alarm file by entering a number\n')
        user_input = input()
        file_index = int(user_input)
        self.break_alarm_file_path = file_list[file_index]

    """
    This method provides a way to run the timer for given number of minutes
    """

    def timer(self, minutes):
            print("\nhit enter to start work timer")
            input()
            start_time = datetime.datetime.now()
            print('\nwork timer:\n')
            work_duration = timedelta(minutes=minutes)
            end_time = start_time + work_duration
            timer_str_check = "00:00"
            sys.stdout.write(timer_str_check)
            while datetime.datetime.now() < end_time:
                timer_seconds = end_time - datetime.datetime.now()
                minutes = timer_seconds.seconds // 60
                seconds = timer_seconds.seconds - (minutes*60)
                if seconds < 10:
                    seconds_str = "0" + str(seconds)
                else:
                    seconds_str = str(seconds)
                if minutes < 10:
                    minutes_str = "0" + str(minutes)
                else:
                    minutes_str = str(minutes)
                timer_str = minutes_str + ":" + seconds_str
                for i in range(0,5):
                    if timer_str_check[i] != timer_str[i]:
                        errase_length = i
                        break
                if errase_length != 5:
                    for i in range(0, 5-errase_length):
                        sys.stdout.write('\b')
                    for i in range(0, 5-errase_length):
                        sys.stdout.write(' ')
                    for i in range(0, 5-errase_length):
                        sys.stdout.write('\b')
                    timer_str_print = timer_str[errase_length:]
                    sys.stdout.write(timer_str_print)
                timer_str_check = timer_str

            #this section runs the alarm for the different operating systems
            if platform.system() == 'Windows':
                #for windows
                os.startfile(self.work_alarm_file_path)
            elif platform.system() == 'Darwin':
                #for mac
                os.system('mpg123 ' + '"' + self.work_alarm_file_path + '"')
                sys.stdout.write('\b\b\b\b\b')
            else:
                #for Linux
                subprocess.call(('xdg-open', self.work_alarm_file_path))

    """
    turns a total minutes into hrs and minutes
    """
    def hrs(self, minutes):
        duration_hrs_str = str(minutes//60)
        if duration_hrs_str == '0':
            duration_hrs_str = ''
        else:
            duration_hrs_str = duration_hrs_str + 'hrs '
        duration_minutes_str = str(minutes - ((minutes//60)*60))
        self.duration_hrs_mins = duration_hrs_str + duration_minutes_str + 'mins'
    """
    finds the length of the sheet. self.sheet_length
    is the next place to enter
    """
    def length(self, sheet):
        sheet_dimensions = sheet.dimensions
        dimension_colon = sheet_dimensions.rfind(":")
        length_start_digit = dimension_colon + 2
        sheet_length = sheet_dimensions[length_start_digit:]
        sheet_length = int(sheet_length)
        if sheet_length < 5:
            self.sheet_length = 5
        else:
            self.sheet_length = sheet_length + 1

    """
    getting a list of unique work_types and time sums
    and printing them
    """
    def totals(self):
        wb = load_workbook(filename=self.filename)
        sheet = wb.active

        self.length(sheet)

        """
        Grabs and prints the list of work-type and duration totals
        """
        work_type_list = []
        work_type_durations_list = []
        for i in range(5, self.sheet_length+1):
            work_type = sheet.cell(row = i, column = 3).value
            if type(work_type) == str:
                work_duration_minutes = sheet.cell(row = i, column = 5).value
                work_type = work_type.lower()
                if work_type not in work_type_list:
                    work_type_list.append(work_type)
                    work_type_durations_list.append(work_duration_minutes)
                else:
                    item_location = work_type_list.index(work_type)
                    work_type_durations_list[item_location] = work_type_durations_list[item_location] + work_duration_minutes
        if len(work_type_list) > 0:
            work_type_list_range = len(work_type_list)
            print('\nproject time sums')
            total_duration_int = 0
            for i in range(0, work_type_list_range):
                duration_int = work_type_durations_list[i]
                total_duration_int = total_duration_int + duration_int
                self.hrs(duration_int)
                print(work_type_list[i] + ': ' + self.duration_hrs_mins)
            self.hrs(total_duration_int)
            print('\ntotal duration for project: ' + self.duration_hrs_mins)

    """
    This sets the work and break lengths
    """
    def info_set(self):
        print('enter number of work minutes or type run:')
        self.workminutes = input()
        try:
            self.workminutes = int(self.workminutes)
        except:
            self.workminutes = 'run'
        print('enter number of break minutes or type run:')
        self.breakminutes = input()
        try:
            self.breakminutes = int(self.breakminutes)
        except:
            self.breakminutes = 'run'
        if type(self.workminutes) and type(self.breakminutes) == int:
            self.entry_work_duration = self.workminutes + self.breakminutes

    """
    This method is for saving data to the spread sheet
    """
    def entry(self):

        wb = load_workbook(filename=self.filename)
        sheet = wb.active

        self.length(sheet)
        sheet_length_str = str(self.sheet_length)

        """
        creating the list of current work-types to choose from and collecting work-type
        """

        #getting a list of unique work_types
        work_type_list = []
        print('\n')
        print('WORK-TYPES:')
        for i in range(5, self.sheet_length+1):
            work_type = sheet.cell(row = i, column = 3).value
            if type(work_type) == str:
                work_type = work_type.lower()
                if work_type not in work_type_list:
                    work_type_list.append(work_type)

        #listing and collecting info
        for x,y in enumerate(work_type_list):
            list_item = str(x) + '. ' + y
            print(list_item)
        print('\nenter work type or a number from the list:')
        work_type_selection = input()
        print('\n')

        #parsing entry
        try:
            work_type_selection_int = int(work_type_selection)
            self.entry_work_type = work_type_list[work_type_selection_int]
        except:
            self.entry_work_type = work_type_selection


        """
        Collecting all other values and saving it to the sheet
        """
        print('\n')
        print('work notes:')
        print(self.work_notes_list)
        print('\nsession notes:')
        self.entry_work_notes = input()
        self.work_notes_list.append(self.entry_work_notes)


        self.entry_date = datetime.datetime.now()
        entry_date = 'B' + sheet_length_str
        sheet[entry_date].value = self.entry_date
        entry_work_type = 'C' + sheet_length_str
        sheet[entry_work_type].value = self.entry_work_type
        entry_work_notes = 'D' + sheet_length_str
        sheet[entry_work_notes].value = self.entry_work_notes
        entry_work_duration = 'E' + sheet_length_str
        sheet[entry_work_duration].value = self.entry_work_duration
        #self.work_duration_position is for adding break minutes later
        #I'm not certain this break minute adding work so I'll have to test
        self.work_duration_position = entry_work_duration
        wb.save(filename=self.filename)
        self.totals()
        print('\ncongratulations on completing a session for: ' + self.entry_work_type)


    """
    This method runs the timers and runs the previous methods, it's the main method
    """
    def work(self):

        self.info_set()

        work_session_on = True
        while work_session_on == True:
            print("\nhit enter to start work timer")
            input()
            start_time = datetime.datetime.now()
            if type(self.workminutes) == int:
                print('\nwork timer:\n')
                if type(self.breakminutes) == int:
                    self.entry_work_duration = self.workminutes + self.breakminutes
                else:
                    self.entry_work_duration = self.workminutes
                work_duration = timedelta(minutes=self.workminutes)
                end_time = start_time + work_duration
                timer_str_check = "00:00"
                sys.stdout.write(timer_str_check)
                while datetime.datetime.now() < end_time:
                    timer_seconds = end_time - datetime.datetime.now()
                    minutes = timer_seconds.seconds // 60
                    seconds = timer_seconds.seconds - (minutes*60)
                    if seconds < 10:
                        seconds_str = "0" + str(seconds)
                    else:
                        seconds_str = str(seconds)
                    if minutes < 10:
                        minutes_str = "0" + str(minutes)
                    else:
                        minutes_str = str(minutes)
                    timer_str = minutes_str + ":" + seconds_str
                    for i in range(0,5):
                        if timer_str_check[i] != timer_str[i]:
                            errase_length = i
                            break
                    if errase_length != 5:
                        for i in range(0, 5-errase_length):
                            sys.stdout.write('\b')
                        for i in range(0, 5-errase_length):
                            sys.stdout.write(' ')
                        for i in range(0, 5-errase_length):
                            sys.stdout.write('\b')
                        timer_str_print = timer_str[errase_length:]
                        sys.stdout.write(timer_str_print)
                    timer_str_check = timer_str

            if self.workminutes == "run":

                timer_str_print = datetime.datetime.now().strftime("%H:%M")
                timer_str_print = "Start Time: " + timer_str_print
                sys.stdout.write(timer_str_print)
                print('\nhit enter when work session is complete')
                input()
                print('\nworking...')
                self.entry_work_duration = datetime.datetime.now() - start_time
                self.entry_work_duration = self.entry_work_duration.seconds // 60


            if platform.system() == 'Windows':
                #for windows
                os.startfile(self.work_alarm_file_path)
            elif platform.system() == 'Darwin':
                #for mac
                os.system('mpg123 ' + '"' + self.work_alarm_file_path + '"')
                sys.stdout.write('\b\b\b\b\b')
            else:
                #for Linux
                subprocess.call(('xdg-open', self.work_alarm_file_path))


            self.entry()

            start_time = datetime.datetime.now()
            if type(self.breakminutes) == int:
                print('\nbreak timer:\n')
                break_duration = timedelta(minutes=self.breakminutes)
                end_time = start_time + break_duration
                timer_str_check = "00:00"
                sys.stdout.write(timer_str_check)
                while datetime.datetime.now() < end_time:
                    timer_seconds = end_time - datetime.datetime.now()
                    minutes = timer_seconds.seconds // 60
                    seconds = timer_seconds.seconds - (minutes*60)
                    if seconds < 10:
                        seconds_str = "0" + str(seconds)
                    else:
                        seconds_str = str(seconds)
                    if minutes < 10:
                        minutes_str = "0" + str(minutes)
                    else:
                        minutes_str = str(minutes)
                    timer_str = minutes_str + ":" + seconds_str
                    for i in range(0,5):
                        if timer_str_check[i] != timer_str[i]:
                            errase_length = i
                            break
                    if errase_length != 5:
                        for i in range(0, 5-errase_length):
                            sys.stdout.write('\b')
                        for i in range(0, 5-errase_length):
                            sys.stdout.write(' ')
                        for i in range(0, 5-errase_length):
                            sys.stdout.write('\b')
                        timer_str_print = timer_str[errase_length:]
                        sys.stdout.write(timer_str_print)
                    timer_str_check = timer_str

            if self.breakminutes == "run":

                timer_str_print = datetime.datetime.now().strftime("%H:%M")
                timer_str_print = "Start Time: " + timer_str_print
                sys.stdout.write(timer_str_print)
                print('\nhit enter when break session is complete')
                input()
                print('\nworking...')
                break_delta = datetime.datetime.now() - start_time
                break_minutes = break_delta.seconds//60
                wb = load_workbook(filename=self.filename)
                sheet = wb.active
                #So I have to save these minutes to goals too,
                #in the same way done in entry
                work_duration_str = sheet[self.work_duration_position].value
                work_duration_int = int(work_duration_str)
                self.entry_work_duration = work_duration_int + break_minutes
                sheet[self.work_duration_position] = self.entry_work_duration
                wb.save(filename=self.filename)

            if platform.system() == 'Windows':
                #for windows
                os.startfile(self.break_alarm_file_path)
            elif platform.system() == 'Darwin':
                #for mac
                os.system('mpg123 ' + '"' + self.break_alarm_file_path + '"')
                sys.stdout.write('\b\b\b\b\b')
            else:
                #for Linux
                subprocess.call(('xdg-open', self.break_alarm_file_path))

    """
    This gives all the stats for the project, including all goals
    (will have to loop through the string lists)
    """

    def stats(self):

        wb = load_workbook(filename=self.filename)
        sheet = wb.active

        self.length(sheet)
        """
        Getting value lists for stats by day
        """
        #list of groups of rows by day
        print('\nWORK DAYS:')
        row_range_by_day_list = []
        current_day = ''
        for i in range(5, self.sheet_length):
            date_value = sheet.cell(row=i, column=2).value
            day_value = date_value.day
            if day_value == current_day:
                row_range_by_day_list[-1].append(i)
            else:
                row_range_by_day_list.append([i])
            current_day = day_value

        #building lists and printing info
        for range_item in row_range_by_day_list:
            work_type_list = []
            work_type_durations_list = []
            for i in range_item:
                #entering the date over and over, should improve this
                day_entry_raw = sheet.cell(row = i, column = 2).value
                day_entry_year = str(day_entry_raw.year)
                day_entry_month = str(day_entry_raw.month)
                day_entry_day = str(day_entry_raw.day)
                day_entry_str = day_entry_year + "/" + day_entry_month + "/" + day_entry_day

                #builds the lists for each range
                work_type = sheet.cell(row = i, column = 3).value
                if type(work_type) == str:
                    work_duration_minutes = sheet.cell(row = i, column = 5).value
                    work_type = work_type.lower()
                    if work_type not in work_type_list:
                        work_type_list.append(work_type)
                        work_type_durations_list.append(work_duration_minutes)
                    else:
                        item_location = work_type_list.index(work_type)
                        work_type_durations_list[item_location] = work_type_durations_list[item_location] + work_duration_minutes

            print('\n' + day_entry_str)
            for i in range(0, len(work_type_list)):
                type_duration_int = work_type_durations_list[i]
                self.hrs(type_duration_int)
                print(work_type_list[i] + ': ' + self.duration_hrs_mins)

        self.totals()
