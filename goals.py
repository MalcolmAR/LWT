import glob, re, datetime, platform
from openpyxl import Workbook, load_workbook

"""
TASK: replace the select project with comma after thing, I keep screwing it up
TASK: make it so it can use a greater range of spreadsheet and music files
"""


class Goals:

    def __init__(self, goals_filename = None):
        self.goals_filename = goals_filename
        self.directory = ''
        self.project_selected = ''
        self.project_goals_list = []
        self.project_list = []
        self.project_list_names = []
        self.goal_list = []
        self.duration_hrs_mins = ''
        self.goal_index_numbers_list = []
        self.goal_times_dict = {}

    """
    This adds notes to the goal. If 'list' is an arg then it just lists notes.
    """
    def notes(self, *args):
        self.glist()
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        print('\nenter the number for the goal you\'re writing notes for:')
        goal_notes_input = input()
        goal_notes_int = int(goal_notes_input)
        goal_notes_cell = sheet.cell(column=7, row=goal_notes_int).value
        new_note_string = False
        notes_strings_list = []
        note_string = ''
        try:
            for x in goal_notes_cell:
                if x == ']':
                     new_note_string = False
                     notes_strings_list.append(note_string)
                     note_string = ''
                elif new_note_string == True:
                     note_string = note_string + x
                elif x == '[':
                     new_note_string = True
            for x in notes_strings_list:
                print('\n' + x)
        except:
            pass
        if 'list' in args:
            pass
        else:
            time_stamp_raw = datetime.datetime.now()
            time_stamp_year = str(time_stamp_raw.year)
            time_stamp_month = str(time_stamp_raw.month)
            time_stamp_day = str(time_stamp_raw.day)
            time_stamp_str = time_stamp_year + "/" + time_stamp_month + "/" + time_stamp_day
            print('\nenter the next note:')
            goal_note_input = input()
            note_string_to_add = '[' + '(' + time_stamp_str + ') ' + goal_note_input + ']'
            if goal_notes_cell:
                goal_notes_cell = goal_notes_cell + note_string_to_add
            else: goal_notes_cell = note_string_to_add
            sheet.cell(column=7, row=goal_notes_int).value = goal_notes_cell
            wb.save(filename=self.goals_filename)

    """
    turns a total minutes into hrs and minutes
    """
    def hrs(self, minutes):
        duration_hrs_str = str(minutes//60)
        if duration_hrs_str == '0':
            duration_hrs_str = ''
        else:
            duration_hrs_str = duration_hrs_str + 'hrs '
        duration_minutes_str = str(minutes - ((minutes//60)*60))
        self.duration_hrs_mins = duration_hrs_str + duration_minutes_str + 'mins'

    """
    finds the length of the sheet. self.sheet_length
    is the next place to enter
    """
    def length(self, sheet):
        sheet_dimensions = sheet.dimensions
        dimension_colon = sheet_dimensions.rfind(":")
        length_start_digit = dimension_colon + 2
        sheet_length = sheet_dimensions[length_start_digit:]
        sheet_length = int(sheet_length)
        if sheet_length < 5:
            self.sheet_length = 5
        else:
            self.sheet_length = sheet_length + 1

    """
    this method lists the goal spread sheets for selection
    """
    def gselect(self):
        file_list = []
        goal_file_list = []
        for file in glob.iglob('**/*.xlsx', recursive=True):
            file_list.append(file)
        for x in file_list:
            if re.search('goals', x, re.IGNORECASE):
                goal_file_list.append(x)
        for key, value in enumerate(goal_file_list):
            print(str(key) + '. ' + value)
        print('\nselect goal file by entering a number\n')
        user_input = input()
        goal_index = int(user_input)
        self.goals_filename = goal_file_list[goal_index]
        #This sets self.directory based on the goal selection
        index = -1
        check = False
        #this if statement adjusts for windows or mac's way of displaying files
        if platform.system() == 'Windows':
            slash_check = '\\'
        else:
            slash_check = '/'
        while check == False:
            if self.goals_filename[index] == slash_check:
                check = True
                self.directory = self.goals_filename[:index+1]
            else:
                index += -1

    """
    This creates a list of goals in the selected file
    The *args is so 'running' can specify running goals only
    This populates self.goal_index_numbers_list, which can be iterated through
    to pull up all relevant goals.
    """
    def glist(self, *args):
        self.goal_index_numbers_list = []
        self.goal_list = []
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        for i in range(5, self.sheet_length):
            if type(sheet.cell(row=i, column=5).value) == datetime.datetime:
                pass
            elif type(sheet.cell(row=i, column=4).value) == datetime.datetime:
                if 'running' in args:
                    pass
                else:
                    self.goal_index_numbers_list.append(i)
                    goal_string = sheet.cell(row=i, column=2).value
                    self.goal_list.append('(completed) ' + goal_string)
            else:
                self.goal_index_numbers_list.append(i)
                goal_string = sheet.cell(row=i, column=2).value
                self.goal_list.append(goal_string)
        if 'lite' in args:
            pass
        else:
            print('\n')
            for key, value in enumerate(self.goal_list):
                key = self.goal_index_numbers_list[key]
                key = str(key)
                print(key + ': ' + value)
            print('\n')


    """
    This creates a list of projects in the directory
    note:
    """
    def plist(self):

        self.project_list = []
        self.project_list_names = []
        for file in glob.iglob(self.directory + '*.*'):
            self.project_list.append(file)
        for key, value in enumerate(self.project_list):
            check = False
            index = -1
            #this if statement adjusts for windows or mac's way of displaying files
            if platform.system() == 'Windows':
                slash_check = '\\'
            else:
                slash_check = '/'
            while check == False:
                if value[index] == slash_check:
                    check = True
                    project_name = value[index+1:-5]
                    self.project_list_names.append(project_name)
                else:
                    index += -1
            key = str(key)
            print(key + ': ' + project_name)

    """
    This lists the projects in the goal sheets directory for
    selection.
    """
    def pselect(self, *args):
        self.plist()
        print('\nselect which project you will edit the goals for by selecting a number')
        project_input = input()
        project_input_int = int(project_input)
        self.project_selected = self.project_list_names[project_input_int]
        #This method creates a list of goals associated with a project
        #and saves it to self.project_goals_list
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        self.project_goals_list = []
        if 'running' in args:
            self.glist('running', 'lite')
        else:
            self.glist('lite')
        try:
            for goal_row in self.goal_index_numbers_list:
                project_list_cell = sheet.cell(row=goal_row, column=6).value
                if self.project_selected in project_list_cell:
                    goal_string = sheet.cell(row=goal_row, column=2).value
                    self.project_goals_list.append((goal_row, goal_string))
        except:
            pass
        #re-setting the following to run stats with a new goal list
        self.goal_index_numbers_list = []
        for key, value in enumerate(self.project_goals_list):
            key = str(key)
            print(key + ': ' + value[1])
            self.goal_index_numbers_list.append(value[0])
        if 'lite' in args:
            self.stats('pselect', 'lite')
        else:
            self.stats('pselect')

    """
    This adds a new goal to the end of the list
    """
    def add(self):
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        print('goal: ')
        goal_input = input()
        print('\n')
        self.plist()
        print('\nCOMMA! select the affiliated project for the goal by entering a number followed by a comma')
        project_selections = input()
        x_index = 0
        digit_string = ''
        project_selection_ints = []
        for x in project_selections:
            if x == ',':
                for i in range(x_index-2, x_index):
                    digit_character = project_selections[i]
                    try:
                        digit_test = int(digit_character)
                        digit_string = digit_string + digit_character
                    except:
                        pass
                try:
                    project_selection_ints.append(digit_string)
                except:
                    pass
                digit_string = ''
            x_index = x_index + 1
        project_list_string = ''
        for x in project_selection_ints:
            x = int(x)
            project_string = self.project_list[x]
            project_list_string =  project_list_string + '[' + project_string + ']'
        sheet.cell(column=6, row=self.sheet_length).value = project_list_string
        sheet.cell(column=2, row=self.sheet_length).value = goal_input
        sheet.cell(column=3, row=self.sheet_length).value = datetime.datetime.now()
        wb.save(filename=self.goals_filename)

    """
    Need to delete a goal, null a goal, edit the goal
    """

    def stop(self):
        self.glist()
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        print('\nenter the number for the goal to stop:')
        goal_stop_input = input()
        goal_stop_int = int(goal_stop_input)
        sheet.cell(column=5, row=goal_stop_int).value = datetime.datetime.now()
        wb.save(filename=self.goals_filename)

    def accomplished(self):
        self.glist('running')
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        print('\nenter the number for the goal accomplished:')
        goal_accomplished_input = input()
        goal_accomplished_int = int(goal_accomplished_input)
        sheet.cell(column=4, row=goal_accomplished_int).value = datetime.datetime.now()
        wb.save(filename=self.goals_filename)
        goal_string = sheet.cell(column=2, row=goal_accomplished_int).value
        print('\ncongratulations on accomplishing the goal: ' + '\n' + goal_string + '\n')

    def edit(self):
        self.glist('running')
        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.length(sheet)
        print('\nenter the number for the goal to edit:')
        goal_edit_input = input()
        goal_edit_int = int(goal_edit_input)
        print('\nwrite out the edited goal:')
        goal_edit_string = input()
        sheet.cell(column=2, row=goal_edit_int).value = goal_edit_string
        wb.save(filename=self.goals_filename)

    """
    this method calculates the total time spent on all goals,
    with breakdowns for each involved project. The argument 'lite'
    will print only the goals and goal time totals (without the project breakdowns).


    """

    def stats(self, *args):

        #this method is important to run first because self.glist populates self.goal_index_numbers_list
        #but pselect() already runs glist so when pselect runs stats it passes
        if 'pselect' in args:
            pass
        else:
            if 'running' in args:
                self.glist('running')
            else:
                self.glist()

        wb = load_workbook(filename=self.goals_filename)
        sheet = wb.active
        self.goal_times_dict = {}
        for goal_row in self.goal_index_numbers_list:
            goal_string = sheet.cell(column=2, row=goal_row).value
            print('\ngoal: ' + goal_string)
            projects_cell = sheet.cell(column=6, row=goal_row).value
            new_project_string = False
            projects_strings_list = []
            project_string = ''
            try:
                for x in projects_cell:
                    if x == ']':
                         new_project_string = False
                         projects_strings_list.append(project_string)
                         project_string = ''
                    elif new_project_string == True:
                         project_string = project_string + x
                    elif x == '[':
                         new_project_string = True
            except:
                pass
            #the following will fill with total times for each work-type in all projects for each goal
            goal_time_by_work =  {}
            for project in projects_strings_list:
                """
                This part sets the first and last row to check in each project
                """
                sum_range_list = []
                sum_wb = load_workbook(filename=project)
                sum_sheet = sum_wb.active
                self.length(sum_sheet)
                #this will get the range of rows in the project that need to be summed-up
                start_found = False
                end_found = False
                for project_row in range(5, self.sheet_length):
                    date_to_check = sum_sheet.cell(column=2, row=project_row).value
                    goal_start_date = sheet.cell(column=3, row=goal_row).value
                    try:
                        if date_to_check > goal_start_date and start_found == False:
                            sum_range_list.append(project_row)
                            start_found = True
                        goal_end_date = sheet.cell(column=4, row=goal_row).value
                        if type(goal_end_date) == datetime.datetime and end_found == False:
                            if date_to_check > goal_end_date:
                                sum_range_list.append(project_row - 1)
                                end_found = True
                    except:
                        pass

                """
                This part builds a list of work-types in each project
                and corresponding total time
                """
                work_type_list = []
                work_type_durations_list = []

                if len(sum_range_list) < 2:
                    end_range = self.sheet_length+1
                else:
                    end_range = sum_range_list[1] + 1
                try:
                    for project_sum_row in range(sum_range_list[0], end_range):
                        work_type = sum_sheet.cell(row = project_sum_row, column = 3).value
                        if type(work_type) == str:
                            work_duration_minutes = sum_sheet.cell(row = project_sum_row, column = 5).value
                            work_type = work_type.lower()
                            if work_type not in work_type_list:
                                work_type_list.append(work_type)
                                work_type_durations_list.append(work_duration_minutes)
                            else:
                                item_location = work_type_list.index(work_type)
                                work_type_durations_list[item_location] = work_type_durations_list[item_location] + work_duration_minutes

                except:
                    pass

                """
                This part prints from the two lists built above,
                and uses them to calculate a total for the project,
                and add to the work-type total for the goal (from all projects)
                """
                total_goal_duration = 0
                if len(work_type_list) > 0:
                    work_type_list_range = len(work_type_list)
                    if 'lite' in args:
                        pass
                    else:
                        print('\nsums for project: \n' + project)
                    total_duration_int = 0
                    for work_type_index in range(0, work_type_list_range):
                        duration_int = work_type_durations_list[work_type_index]
                        total_duration_int = total_duration_int + duration_int
                        self.hrs(duration_int)
                        work_type_string = work_type_list[work_type_index]
                        if 'lite' in args:
                            pass
                        else:
                            print(work_type_string + ': ' + self.duration_hrs_mins)
                        if work_type_string in goal_time_by_work:
                            goal_time_by_work[work_type_string] += duration_int
                        else:
                            goal_time_by_work[work_type_string] = duration_int
                    total_goal_duration += total_duration_int
                    self.hrs(total_duration_int)
                    if 'lite' in args:
                        pass
                    else:
                        print('total duration for project: ' + self.duration_hrs_mins)
            if 'lite' in args:
                pass
            else:
                print('\nwork-type totals for the goal: \n')
                for key in goal_time_by_work:
                    time_int = goal_time_by_work[key]
                    self.hrs(time_int)
                    print(key + ': ' + self.duration_hrs_mins)
            try:
                self.hrs(total_goal_duration)
                self.goal_times_dict[goal_string] = self.duration_hrs_mins
                print('\ntotal duration for the goal: ' + self.duration_hrs_mins)
            except:
                pass
